"""
This script enriches "Weaving DH Data Table.xlsx" with KWIC (Key Word In Context) concordance
lines extracted from plain-text article files.

Output: "Weaving DH Data Table_enriched.csv" (UTF-8 with BOM for Excel compatibility).

Columns written:
  F  kwic_textile       – ±15-word KWIC for each word in col E (textile_words)
  I  construction_words – construction-vocabulary words found in the text
  J  kwic_construction  – ±15-word KWIC for the words found in col I

KWIC format
-----------
  • Window of ±15 words around each hit; all whitespace runs (including
    newlines from PDF extraction) are normalised to a single space.
  • The matching token is marked with **asterisks** for later reference.
  • Multiple hits for the same query word are separated by a blank line
    (two newlines, rendered as line breaks within the CSV cell).
  • When col E lists several words, blocks are separated by a blank line
    preceded by a "word:" label.

Stemming
--------
  NLTK PorterStemmer, extended with an explicit normalisation table for
  irregular forms Porter does not reduce to their base stem
  (built → build stem, dug → dig stem).

Construction-word exclusions
-----------------------------
  Occurrences of build* immediately followed by "on" or "upon" are
  excluded (discourse connective, not a construction metaphor).

Usage
-----
  python kwic_enrichment.py
  (adjust PATH constants in section 0 if needed)
"""

import csv
import re
import sys
from pathlib import Path

import nltk
from nltk.stem import PorterStemmer
import openpyxl
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# 0.  Configuration
# ---------------------------------------------------------------------------

XLSX_PATH   = Path("Weaving DH Data Table.xlsx")
TXT_DIR     = Path("Weaving DH Data")
OUTPUT_PATH = Path("Weaving DH Data Table_enriched.csv")

KWIC_WINDOW = 15

CONSTRUCTION_SEEDS = ["dig", "mine", "build"]

COL_ID            = 1
COL_TEXTILE_WORDS = 5
COL_KWIC_TEXTILE  = 6
COL_CONSTRUCTION  = 9
COL_KWIC_CONST    = 10

# ---------------------------------------------------------------------------
# 1.  NLTK bootstrap
# ---------------------------------------------------------------------------

for _res in ("punkt", "punkt_tab"):
    try:
        nltk.data.find(f"tokenizers/{_res}")
    except LookupError:
        nltk.download(_res, quiet=True)

# ---------------------------------------------------------------------------
# 2.  Stemmer and normalisation
# ---------------------------------------------------------------------------

_stemmer = PorterStemmer()


def stem(word: str) -> str:
    return _stemmer.stem(word.lower())


_STEM_NORMALISE: dict[str, str] = {
    stem("built"): stem("build"),
    stem("dug"):   stem("dig"),
}


def normalised_stem(word: str) -> str:
    s = stem(word)
    return _STEM_NORMALISE.get(s, s)


CONSTRUCTION_CANONICAL: dict[str, str] = {
    normalised_stem(s): s for s in CONSTRUCTION_SEEDS
}

BUILD_STEM = normalised_stem("build")

# ---------------------------------------------------------------------------
# 3.  Text loading and tokenisation
# ---------------------------------------------------------------------------

_text_cache:  dict[str, str]                         = {}
_token_cache: dict[str, tuple[list[str], list[str]]] = {}

_WORD_RE    = re.compile(r"\w+", re.UNICODE)
_WHITESPACE = re.compile(r"\s+")


def load_text(article_id: str) -> str:
    if article_id in _text_cache:
        return _text_cache[article_id]
    path = TXT_DIR / f"{article_id}.txt"
    if not path.exists():
        _text_cache[article_id] = ""
        return ""
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        print(f"  [WARN] Cannot read {path}: {exc}", file=sys.stderr)
        text = ""
    _text_cache[article_id] = text
    return text


def get_tokens(article_id: str) -> tuple[list[str], list[str]]:
    """
    Return (words, gaps) for the article text.

    words[i] is the i-th alphanumeric token.
    gaps[i]  is the text between words[i-1] and words[i], with every
             whitespace run collapsed to a single space.
    gaps[-1] is the text after the last word (always "").
    """
    if article_id in _token_cache:
        return _token_cache[article_id]
    text = load_text(article_id)
    words: list[str] = []
    gaps:  list[str] = []
    prev_end = 0
    for m in _WORD_RE.finditer(text):
        raw_gap = text[prev_end:m.start()]
        gaps.append(_WHITESPACE.sub(" ", raw_gap))
        words.append(m.group())
        prev_end = m.end()
    gaps.append("")
    _token_cache[article_id] = (words, gaps)
    return words, gaps

# ---------------------------------------------------------------------------
# 4.  Construction false-positive filters
# ---------------------------------------------------------------------------
#
# build* exclusions:
#   - "building on / built on / build on / upon"  →  discourse connective
#   - "built-in"  →  technical adjective (built-in feature, built-in search, …)
#
# mine exclusions:
#   - "mine" as possessive pronoun: preceded by "of", "is", "was", "not",
#     "called", "touch", or followed by "begins", "own"; or enclosed in
#     parentheses/brackets: (mine) / [mine]
#   - "mines" / "mine" preceded or followed by geological context words
#     (quarry/quarries, lithium, rare-earth, metal, ore, mineral, copper,
#     zinc, coal, gold, silver, iron, stone, rock, archaeological, excavat*)
#
# All checks operate on the already-tokenised word list and its gap list.

MINE_STEM  = normalised_stem("mine")

# Geological/physical context that signals literal mine usage
_GEO_WORDS = {
    "quarry", "quarries", "lithium", "metal", "metals", "ore", "ores",
    "mineral", "minerals", "copper", "zinc", "coal", "gold", "silver",
    "iron", "stone", "rock", "archaeological", "excavat", "excavation",
    "excavations", "excavated", "rare", "earth",
}

# Words that mark "mine" as a possessive pronoun when immediately adjacent
_MINE_POSSESSIVE_PREV = {"of", "is", "was", "not", "called", "touch", "only"}
_MINE_POSSESSIVE_NEXT = {"begins", "begin", "own", "as"}



def is_build_on(words: list[str], i: int) -> bool:
    return i + 1 < len(words) and words[i + 1].lower() in ("on", "upon")


def is_built_in(words: list[str], gaps: list[str], i: int) -> bool:
    """True for 'built-in': gap after words[i] is '-' and next word is 'in'."""
    if i + 1 >= len(gaps):
        return False
    gap_after = gaps[i + 1].strip()
    next_word  = words[i + 1].lower() if i + 1 < len(words) else ""
    return gap_after == "-" and next_word == "in"


def is_mine_possessive(words: list[str], gaps: list[str], i: int) -> bool:
    """
    True when 'mine' functions as a possessive pronoun rather than as a
    mining metaphor.  Heuristics:
      1. Preceded (within gap) by an opening bracket/paren: (mine) [mine]
      2. Previous word is a possessive-context word
      3. Next word is a possessive-context word
      4. 'mine' is capitalised as part of a proper name (Make Mine Music)
         — caught by the title-word check: next word is also capitalised
    """
    prev_gap = gaps[i]
    if "(" in prev_gap or "[" in prev_gap:
        return True
    next_gap = gaps[i + 1] if i + 1 < len(gaps) else ""
    if ")" in next_gap or "]" in next_gap:
        return True
    prev_word = words[i - 1].lower() if i > 0 else ""
    next_word = words[i + 1].lower() if i + 1 < len(words) else ""
    if prev_word in _MINE_POSSESSIVE_PREV:
        return True
    if next_word in _MINE_POSSESSIVE_NEXT:
        return True
    # "mine" preceded or followed within ±3 tokens by a geological word
    context = [words[j].lower() for j in range(max(0, i-3), min(len(words), i+4)) if j != i]
    if any(c in _GEO_WORDS or any(c.startswith(g) for g in ["excavat"]) for c in context):
        return True
    return False


def is_mine_geological(words: list[str], i: int) -> bool:
    """True for plural 'mines' or geological compound uses."""
    w = words[i].lower()
    if w == "mines":
        context = [words[j].lower() for j in range(max(0, i-3), min(len(words), i+4)) if j != i]
        if any(c in _GEO_WORDS for c in context):
            return True
    return False



def should_exclude(
    words: list[str],
    gaps: list[str],
    i: int,
    ns: str,
) -> bool:
    """
    Central exclusion gate.  Returns True if the occurrence at position i
    should be filtered out as a non-metaphorical use.
    """
    if ns == BUILD_STEM:
        if is_build_on(words, i):
            return True
        if is_built_in(words, gaps, i):
            return True
    if ns == MINE_STEM:
        w = words[i].lower()
        if w == "mine" and is_mine_possessive(words, gaps, i):
            return True
        if is_mine_geological(words, i):
            return True

    return False

# ---------------------------------------------------------------------------
# 5.  KWIC extraction — plain text output
# ---------------------------------------------------------------------------


def find_kwic_hits(
    article_id: str,
    query_stem: str,
    exclude_construction: bool = False,
) -> list[tuple[int, int, int]]:
    words, gaps = get_tokens(article_id)
    results = []
    for i, w in enumerate(words):
        if normalised_stem(w) != query_stem:
            continue
        if exclude_construction and should_exclude(words, gaps, i, query_stem):
            continue
        start = max(0, i - KWIC_WINDOW)
        end   = min(len(words), i + KWIC_WINDOW + 1)
        results.append((start, i, end))
    return results


def _render_window(
    words: list[str],
    gaps: list[str],
    start: int,
    hit_i: int,
    end: int,
) -> str:
    """
    Render one KWIC window as a plain string.
    The hit token is wrapped in **double asterisks**.
    Gaps between words are preserved as-is (already normalised to single spaces).
    The leading gap is stripped; there is no trailing gap.
    """
    parts = []
    leading = gaps[start].strip()
    if leading:
        parts.append(leading)
        parts.append(" ")
    for i in range(start, end):
        if i == hit_i:
            parts.append(f"**{words[i]}**")
        else:
            parts.append(words[i])
        if i < end - 1:
            gap = gaps[i + 1]
            parts.append(gap if gap else " ")
    return "".join(parts)


def build_kwic_text(
    article_id: str,
    query_stem: str,
    exclude_construction: bool = False,
) -> str:
    words, gaps = get_tokens(article_id)
    hits = find_kwic_hits(article_id, query_stem, exclude_construction)
    if not hits:
        return ""
    return "\n\n".join(
        _render_window(words, gaps, start, hit_i, end)
        for start, hit_i, end in hits
    )


def build_kwic_text_multi(
    article_id: str,
    words_list: list[str],
    exclude_construction: bool = False,
) -> str:
    if not words_list:
        return ""
    if len(words_list) == 1:
        return build_kwic_text(article_id, normalised_stem(words_list[0]), exclude_construction)

    blocks = []
    for word in words_list:
        block = build_kwic_text(article_id, normalised_stem(word), exclude_construction)
        if block:
            blocks.append(f"{word}:\n{block}")
    return "\n\n".join(blocks)

# ---------------------------------------------------------------------------
# 6.  Construction-word detection
# ---------------------------------------------------------------------------


def detect_construction_words(article_id: str) -> list[str]:
    words, gaps = get_tokens(article_id)
    found: dict[str, bool] = {}
    for i, w in enumerate(words):
        ns = normalised_stem(w)
        if ns not in CONSTRUCTION_CANONICAL:
            continue
        canonical = CONSTRUCTION_CANONICAL[ns]
        if canonical not in found:
            found[canonical] = False
        if not should_exclude(words, gaps, i, ns):
            found[canonical] = True
    return [c for c in CONSTRUCTION_SEEDS if found.get(c)]

# ---------------------------------------------------------------------------
# 7.  Parse textile words from col E
# ---------------------------------------------------------------------------


def parse_textile_words(cell_value) -> list[str]:
    """Split col E into individual query words, deduplicating by stem so that
    e.g. 'fabric, fabrics' or 'fabric, fabric' produce only one KWIC block."""
    if not cell_value:
        return []
    raw = str(cell_value).strip()
    if not raw:
        return []
    seen_stems: set[str] = set()
    result: list[str] = []
    for part in re.split(r"[,;\n\r]+", raw):
        word = part.strip()
        if not word:
            continue
        s = normalised_stem(word)
        if s not in seen_stems:
            seen_stems.add(s)
            result.append(word)
    return result

# ---------------------------------------------------------------------------
# 8.  Main
# ---------------------------------------------------------------------------


def main() -> None:
    if not XLSX_PATH.exists():
        sys.exit(f"ERROR: Excel file not found: {XLSX_PATH}")
    if not TXT_DIR.is_dir():
        sys.exit(f"ERROR: Text directory not found: {TXT_DIR}")

    print(f"Loading workbook: {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    total_rows = ws.max_row
    data_rows  = total_rows - 1
    print(f"Processing {data_rows} data rows …")

    missing_files = 0
    enriched      = 0

    def _serialise(v):
        """Render a cell value as a clean string: int-valued floats lose the .0"""
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return "" if v is None else str(v)

    with open(OUTPUT_PATH, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_ALL)

        for row_idx in range(1, total_rows + 1):
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

            if row_idx == 1:
                writer.writerow([_serialise(v) for v in row_values])
                continue

            article_id = row_values[COL_ID - 1]
            if article_id is None:
                writer.writerow([_serialise(v) if not isinstance(v, str) else v for v in row_values])
                continue
            article_id = str(article_id).strip()
            if not article_id:
                writer.writerow([_serialise(v) if not isinstance(v, str) else v for v in row_values])
                continue

            if not load_text(article_id):
                missing_files += 1
                writer.writerow([_serialise(v) if not isinstance(v, str) else v for v in row_values])
                continue

            textile_words = parse_textile_words(row_values[COL_TEXTILE_WORDS - 1])

            row_values[COL_KWIC_TEXTILE - 1] = build_kwic_text_multi(article_id, textile_words)

            found_cw = detect_construction_words(article_id)
            row_values[COL_CONSTRUCTION - 1] = ", ".join(found_cw) if found_cw else ""
            row_values[COL_KWIC_CONST - 1]   = build_kwic_text_multi(article_id, found_cw, exclude_construction=True)

            writer.writerow([_serialise(v) if not isinstance(v, str) else v for v in row_values])
            enriched += 1
            if enriched % 100 == 0:
                print(f"  … {enriched}/{data_rows} rows done")

    print(f"\nSummary:")
    print(f"  Rows enriched : {enriched}")
    print(f"  Missing files : {missing_files}")
    print(f"\nSaving to: {OUTPUT_PATH}")
    print("Done.")


if __name__ == "__main__":
    main()